from django.shortcuts import render, redirect, get_object_or_404
from .models import Articulos, SolicitudReserva, Cliente
from .forms import ArticulosForm, CustomUserForm, SolicitudReservaForm, SolicitudReservaConComboBoxForm
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth import login, authenticate
from django.db.models import F
from django.views.generic import ListView, View
from django.http import HttpResponse
from .utils import render_to_pdf
from django.contrib import messages
from django.views.generic.base import TemplateView
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference, series, LineChart
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from itertools import chain
from rest_framework import viewsets
from .serializers import ArticulosSerializers, SolicitudReservaSerializers
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.utils import get_column_letter
import pandas as pd
from django.http import FileResponse
import xlsxwriter
from io import BytesIO
from django.shortcuts import render
from django.http import HttpResponse
from django_pandas.io import read_frame
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.marker import DataPoint




def inicio(request):

    return render(request, 'core/inicio.html')

def login(request):
    
    return render(request, 'core/login.html')


def solicitarreserva(request):
    data = {
        'form' : SolicitudReservaConComboBoxForm()
    }

    if request.method == 'POST':
        form= SolicitudReservaConComboBoxForm(request.POST)
        if form.is_valid():
            
            
            messages.success(request, "Solicitud registrada")
            data['mensaje'] = "Enviado correctamente"
            return redirect(to="listadosolicitudes")
            form.save()
        pass
    else:
        form = SolicitudReservaConComboBoxForm()   
    return render(request, 'core/solicitudreserva.html', {'form': form})

def ver_articulo(request, id_articulos):
    articulos = get_object_or_404(Articulos, id=id_articulos)
    return render(request, 'ver_articulo.html', {'articulo': articulos})

def listadoarticulos(request):
    articulos = Articulos.objects.all()
    data = {
        'articulos':articulos
    }
    return render(request, 'core/listadoarticulos.html', data)

def listadosolicitudes(request):
    solicitudes = SolicitudReserva.objects.all()
    data = {
        'solicitudes': solicitudes
    }
    return render(request, 'core/listadosolicitudes.html', data)


@login_required

def nuevoarticulo(request):

    data = {
        'form' : ArticulosForm()
    }

    if request.method == 'POST':
        formulario = ArticulosForm(request.POST, files=request.FILES)
        if formulario.is_valid():
            formulario.save()
            messages.success(request, "Articulo registrado")
            data['mensaje'] = "Guardado correctamente"
            return redirect(to="listadoarticulos")

    return render(request, 'core/nuevoarticulo.html', data)

@permission_required('core.change_articulos')
def modificararticulo(request, id):
    articulos = Articulos.objects.get(id=id)
    data = {
        'form':ArticulosForm(instance=articulos)
    }

    if request.method == 'POST':
        formulario = ArticulosForm(data=request.POST, instance=articulos, files=request.FILES)
        if formulario.is_valid():
            formulario.save()
            messages.success(request, "Modificado correctamente")
            data['mensaje'] = "Modificado correctamente"
            data ["form"] = ArticulosForm(instance=Articulos.objects.get(id=id))
            return redirect(to="listadoarticulos")

    return render(request, 'core/modificararticulo.html', data)
    
@permission_required('core.delete_articulos')
def eliminararticulo(request, id):
    articulo = Articulos.objects.get(id=id)
    articulo.delete()
    messages.success(request, "Eliminado correctamente")
    return redirect(to="listadoarticulos")

def registrousuario(request):
    formulario = CustomUserForm(request.POST or None )
    data = {
        'form':formulario
    }

    if formulario.is_valid():
            formulario.save()
            return redirect(to='inicio')

    #if request.method == 'POST':
        #formulario = CustomUserForm(request.POST or None )
        

            #username = formulario.cleaned_data['username']
            #password = formulario.cleaned_data['password1']
            #user = authenticate(username=username, password=password)
            #login(request, user)
            

    return render(request, 'registration/registrar.html', data)



class ArticulosViewSet(viewsets.ModelViewSet):
    queryset = Articulos.objects.all()
    serializer_class = ArticulosSerializers

class SolicitudReservaViewSet(viewsets.ModelViewSet):
    queryset = SolicitudReserva.objects.all()
    serializer_class = SolicitudReservaSerializers


@permission_required('core.change_articulos')
def modificarsolicitud(request, id):
    solicitudes = SolicitudReserva.objects.get(id=id)
    data = {
        'form':SolicitudReservaForm(instance=solicitudes)
    }

    if request.method == 'POST':
        formulario = SolicitudReservaForm(data=request.POST, instance=solicitudes, files=request.FILES)
        if formulario.is_valid():
            formulario.save()
            messages.success(request, "Modificado correctamente")
            data['mensaje'] = "Modificado correctamente"
            data ["form"] = SolicitudReservaForm(instance=SolicitudReserva.objects.get(id=id))
            return redirect(to="listadosolicitudes")

    return render(request, 'core/modificarsolicitud.html', data)

@permission_required('core.delete_SolicitudReserva')
def eliminarsolicitud(request, id):
    solicitudes = SolicitudReserva.objects.get(id=id)
    solicitudes.delete()
    messages.success(request, "Eliminado correctamente")
    return redirect(to="listadosolicitudes")

class ListArticulosPdf(View):
    def get(self, request, *args, **kwargs):
        articulos = Articulos.objects.all()
        data = {
            'articulos': articulos
        }
        pdf = render_to_pdf("core/reportearticulos.html", data)
        return HttpResponse(pdf, content_type='application/pdf')

class ListSolicitudesPdf(View):
    def get(self, request, *args, **kwargs):
        solicitudes = SolicitudReserva.objects.all()
        data = {
            'solicitudes': solicitudes
        }
        pdf = render_to_pdf("core/reportesolicitudes.html", data)
        return HttpResponse(pdf, content_type='application/pdf')

#def update_stock(sender, instance, **kwargs):
    #instance.articulos.stock_sede -= instance.cantidad_solicitada
    #instance.articulos.save()

# register the signal
#signals.post_save.connect(update_stock, sender=SolicitudReserva, dispatch_uid="update_stock_count")

def ReportesExcel(request):
     # Crear un libro de Excel y agregar una hoja de cálculo
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Articulos'

    # Obtener los datos de la base de datos y agregarlos a la hoja de cálculo
    articulos = Articulos.objects.all()
    sheet.column_dimensions['A'].width = 25
    sheet.column_dimensions['B'].width = 25
    sheet.column_dimensions['C'].width = 25
    sheet['A5'] = 'Nombre'
    sheet['B5'] = 'Cantidad'
    row_num = 6
    
    for articulo in articulos:
        sheet.cell(row=row_num, column=1, value=articulo.nombre_articulo).alignment = Alignment(horizontal = "center")
        sheet.cell(row=row_num, column=1, value=articulo.nombre_articulo).border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                                                                         top=Side(border_style="thin"), bottom=Side(border_style="thin"))

        sheet.cell(row=row_num, column=2, value=articulo.stock_sede).alignment = Alignment(horizontal = "center")
        sheet.cell(row=row_num, column=2, value=articulo.stock_sede).border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                                                                               top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        row_num += 1
    #Decorando el título
    sheet['B1'].alignment = Alignment(horizontal = "center",vertical = "center")
    sheet['B1'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                              top = Side(border_style = "thin"), bottom = Side(border_style = "thin") ) 
    sheet['B1'].fill = PatternFill(start_color = '66FFCC', end_color = '66FFCC', fill_type = "solid")
    sheet['B1'].font = Font(name = 'Calibri', size = 12, bold = True)
    sheet['B1'] = 'REPORTE DE ARTICULOS'
    sheet.merge_cells('B1:E1')
    #Decorando la tabla
    sheet['A5'].alignment = Alignment(horizontal = "center",vertical = "center")
    sheet['A5'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                              top = Side(border_style = "thin"), bottom = Side(border_style = "thin") ) 
    sheet['A5'].fill = PatternFill(start_color = '66FFCC', end_color = '66FFCC', fill_type = "solid")
    sheet['A5'].font = Font(name = 'Calibri', size = 12, bold = True)

    sheet['B5'].alignment = Alignment(horizontal = "center",vertical = "center")
    sheet['B5'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                              top = Side(border_style = "thin"), bottom = Side(border_style = "thin") ) 
    sheet['B5'].fill = PatternFill(start_color = '66FFCC', end_color = '66FFCC', fill_type = "solid")
    sheet['B5'].font = Font(name = 'Calibri', size = 12, bold = True)
   
    # Crear un gráfico de barras y agregarlo a la hoja de cálculo
    chart = BarChart()
    chart.title = 'Inventario de artículos'
    chart.y_axis.title = 'Cantidad'
    chart.x_axis.title = 'Artículo'
    data = Reference(sheet, min_col=2, min_row=1, max_row=row_num-1, max_col=2)
    categories = Reference(sheet, min_col=1, min_row=2, max_row=row_num-1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    sheet.add_chart(chart, "E5")

    # Configurar la respuesta HTTP con el archivo de Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=articulos.xlsx'
    workbook.save(response)

    return response


def SolicitudesExcel(request):
    workbook = openpyxl.Workbook()
    sheet1 = workbook.active
    sheet1.title = 'Reservas'

    # Obtener los datos de la base de datos y agregarlos a la hoja de cálculo
    reserva = SolicitudReserva.objects.all()
    sheet1.column_dimensions['A'].width = 25
    sheet1.column_dimensions['B'].width = 25
    sheet1.column_dimensions['C'].width = 25
    sheet1['A5'] = 'Nombre del Equipo'
    sheet1['B5'] = 'Cantidad Solicitada'
    sheet1['C5'] = 'Nombre del solicitante'
    row_num = 6
    for reservas in reserva:
        sheet1.cell(row=row_num, column=1, value=reservas.nombre_equipo).alignment = Alignment(horizontal = "center")
        sheet1.cell(row=row_num, column=1, value=reservas.nombre_equipo).border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                                                                         top=Side(border_style="thin"), bottom=Side(border_style="thin"))

        sheet1.cell(row=row_num, column=2, value=reservas.cantidad_solicitada).alignment = Alignment(horizontal = "center")
        sheet1.cell(row=row_num, column=2, value=reservas.cantidad_solicitada).border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                                                                               top=Side(border_style="thin"), bottom=Side(border_style="thin"))

        sheet1.cell(row=row_num, column=3, value=reservas.nombre).alignment = Alignment(horizontal = "center")
        sheet1.cell(row=row_num, column=3, value=reservas.nombre).border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                                                                  top=Side(border_style="thin"), bottom=Side(border_style="thin"))

        row_num += 1
    #Decorando el título
    sheet1['B1'].alignment = Alignment(horizontal = "center",vertical = "center")
    sheet1['B1'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                              top = Side(border_style = "thin"), bottom = Side(border_style = "thin") ) 
    sheet1['B1'].fill = PatternFill(start_color = '66FFCC', end_color = '66FFCC', fill_type = "solid")
    sheet1['B1'].font = Font(name = 'Calibri', size = 12, bold = True)
    sheet1['B1'] = 'REPORTE DE RESERVAS'
    sheet1.merge_cells('B1:E1')
    #Decorando la tabla
    sheet1['A5'].alignment = Alignment(horizontal = "center",vertical = "center")
    sheet1['A5'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                              top = Side(border_style = "thin"), bottom = Side(border_style = "thin") ) 
    sheet1['A5'].fill = PatternFill(start_color = '66FFCC', end_color = '66FFCC', fill_type = "solid")
    sheet1['A5'].font = Font(name = 'Calibri', size = 14, bold = True)

    sheet1['B5'].alignment = Alignment(horizontal = "center",vertical = "center")
    sheet1['B5'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                              top = Side(border_style = "thin"), bottom = Side(border_style = "thin") ) 
    sheet1['B5'].fill = PatternFill(start_color = '66FFCC', end_color = '66FFCC', fill_type = "solid")
    sheet1['B5'].font = Font(name = 'Calibri', size = 14, bold = True)

    sheet1['C5'].alignment = Alignment(horizontal = "center",vertical = "center")
    sheet1['C5'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                              top = Side(border_style = "thin"), bottom = Side(border_style = "thin") ) 
    sheet1['C5'].fill = PatternFill(start_color = '66FFCC', end_color = '66FFCC', fill_type = "solid")
    sheet1['C5'].font = Font(name = 'Calibri', size = 14, bold = True)
    
    # Crear un gráfico de barras y agregarlo a la hoja de cálculo
    chart = BarChart()
    chart.title = 'SOLICITUDES DE RESERVA'
    chart.y_axis.title = 'Cantidad'
    chart.x_axis.title = 'Artículo'
    data = Reference(sheet1, min_col=2, min_row=1, max_row=row_num-1, max_col=2)
    categories = Reference(sheet1, min_col=1, min_row=2, max_row=row_num-1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    sheet1.add_chart(chart, "E5")
    # Configurar la respuesta HTTP con el archivo de Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=solicitudes.xlsx'
    workbook.save(response)

    return response
